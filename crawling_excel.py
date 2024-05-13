from selenium import webdriver
import chromedriver_autoinstaller
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time, os
from datetime import datetime
from openpyxl import Workbook, load_workbook

now = datetime.now().strftime(f"%Y-%m-%d")

# 크롤링 함수
def danawa_crawling(search_query):
    options = Options()
    options.add_argument('headless')  # 브라우저 창 안보이기
    chromedriver_autoinstaller.install()
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(5)  # 크롬 브라우저 내부 대기
    driver.set_window_size(1920, 1280)  # 브라우저 사이즈
    # 검색 URL 업데이트
    base_url = "https://search.danawa.com/dsearch.php"
    driver.get(f"{base_url}?query={search_query}")
    # 2초 대기, 페이지 로드 완료를 기다립니다.
    time.sleep(2)
    # BeautifulSoup을 이용한 HTML 파싱
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    # 상품 리스트 선택
    goods_list = soup.select("#productListArea > div.main_prodlist.main_prodlist_list>ul>li>div.prod_main_info")

    new_list=[]
    for v in goods_list:
        name=""
        price=""
        url=""
        if v.find('div', class_='prod_info'):
            name = v.select_one('p.prod_name > a').text.strip()
            url=v.select_one('p.prod_name > a').get('href')
        if v.find('div', class_='prod_pricelist'):
            price = v.select_one('ul>li>p> a').text.strip()
            price = price.replace(',', '').replace('원', '')
        if(url.startswith("https")):
            #print(f"이름 : {name}, 가격 : {price}, 주소 : {url}")
            #driver.close()
            new_list.append([name,price,url])

    driver.close()
    return new_list

# 엑셀에 저장하는 함수
def save_to_excel(data_list, filename):
    file_path = f"database\\{filename}"

    #만약에 엑셀 파일이 존재하면 있는 파일 가져오기, 그렇지 않으면 새 엑셀파일 생성하기
    if os.path.isfile(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["검색어", "제품명", "가격", "링크", "id", "app"])  # 엑셀 파일의 헤더 생성
    # for row in data_list:
    #     ws.append(row)

    # 검색한 제품의 최저값만 엑셀에 저장하도록 변경
    ws.append(data_list)
    wb.save(file_path)
    print(f"데이터가 '{filename}'에 저장되었습니다.")


# if _name_ == "_main_":
#     list = danawa_crawling("노트북")
#     save_to_excel(list, f"{now}_product_info.xlsx")

# list2=danawa_crawling("아이폰 15 프로")
# print(list2)

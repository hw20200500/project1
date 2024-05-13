import time, threading, os, time
import send_message
import crawling_excel
from flask import Flask, render_template, request, redirect
import asyncio
from openpyxl import load_workbook



app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    product = request.form['search']
    alarm = request.form['alarm']
    id = request.form['id']
    data = crawling_excel.danawa_crawling(product)
    product_list = data[:5]

    # database 변수는 엑셀 파일에 저장할 값들을 저장하는 변수.
    # 검색어, 최저가 상품명, 가격, 링크, 사용자 id, 채팅앱(app) 순으로 저장
    # 채팅에서 출력하는 product_list를 가격을 기준으로 정렬하여 가장 싼 데이터만 저장.
    sorted_prices = sorted(product_list, key=lambda x: int(x[1]))
    database = sorted_prices[0]
    database.append(id)
    database.append(alarm)
    database.insert(0, product)
    
    crawling_excel.save_to_excel(database, f"product_info.xlsx")

    # 채팅봇이 보낼 메시지 내용 생성
    message  = f"{product}의 최저가 상품 리스트입니다.\n\n"

    # 찾은 제품들 중 유사도가 높은 제품 5개만 추출해서 출력
    for i in range(5):
        for text in sorted_prices[i]:
            message = message+text+"\n"
        message = message+"="*50+"\n\n"
    
    # 터미널에서 메시지 내용 확인하기
    print(message)

    # 선택한 알람이 ntfy일 경우,
    if alarm=="ntfy":
        # 사용자마다 다른 ntfy 채팅방을 생성하기 위해서 사용자 id별로 다른 채팅방 생성 및 그곳에 메시지 전송
        send_message.ntfy0(f"danawa_{id}", message)

        # 없는 채팅방(ntfy 링크)을 새로 생성할 경우 시간이 좀 소요되므로 10초정도 뒤에 ntfy 채팅방 링크로 이동
        time.sleep(10)
        return redirect(f"https://ntfy.sh/danawa_{id}")
    elif alarm=="slack":
        state = send_message.sendSlackWebHook(message)
        
        # 아직 코드를 못받아서 그냥 메시지를 화면에 출력하는 것으로 임시 설정.
        return state
        # return redirect("https://app.slack.com/client/T072KG05545/C072QT9TFJQ")
    else:
        # 텔레그램 실행. 아직 테스트 안해봄.
        asyncio.run(send_message.telegram_bot(id, message))
    

# 스레드를 이용하여 데이터베이스에 저장되어 있는 제품들 중 더 저렴한 제품이 있는지 없는지 확인하는 함수
def scheduled_task(interval_minutes):
    while True:
        # 특정 함수를 호출합니다.
        file_path = f"database\\product_info.xlsx"
        if os.path.isfile(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                # 맨 처음 칼럼인 row를 제외한 나머지 row에서 실행하기
                if row!=('검색어', '제품명', '가격', '링크', 'id', 'app'):
                    
                    print(f"이전에 {row[-2]}님이 검색한 제품 {row[0]}에 대한 최저값 검색을 시작합니다.")
                    products_list = crawling_excel.danawa_crawling(row[0])
                    products_list = products_list[:5]
                    sorted_prices = sorted(products_list, key=lambda x: int(x[1]))
                    # 지금 찾은 제품의 가격이 엑셀에 저장된 동일 제품의 가격보다 더 저렴한 경우 실행
                    if int(sorted_prices[0][1])<int(row[2]):

                        # 엑셀에 저장되어 있는 제품 데이터 업데이트
                        for i in range(1, 4):
                            ws.cell(row=idx, column=i).value = sorted_prices[0][i-1]
                        
                        # 알림 메시지 작성
                        message = f"이전에 검색하신 {row[0]} 상품보다 더 저렴한 제품을 발견했습니다.\n\n"
                        for text in sorted_prices[0]:
                            message = message+text+"\n"
                        print(message)

                        # 엑셀에 저장된 app이 ntfy일 때
                        if row[-1]=="ntfy":
                            send_message.ntfy0(f"danawa_{row[-2]}", message)

                        # 엑셀에 저장된 app이 slack일 때
                        elif row[-1]=="slack":
                            send_message.sendSlackWebHook(message)
                            # return redirect("https://app.slack.com/client/T072KG05545/C072QT9TFJQ")
                        
                        # 엑셀에 저장된 app이 텔레그램일 때
                        else:
                            asyncio.run(send_message.telegram_bot(row[-2], message))

                    else : print("해당 제품의 더 싼 제품이 아직 없습니다.")
                
        # 주어진 시간 간격 동안 대기합니다.
        time.sleep(interval_minutes * 60)  # 분을 초로 변환하여 사용합니다


# 스레드를 생성하여 예약된 작업을 실행합니다.
thread = threading.Thread(target=scheduled_task, args=(1,))  # 1분마다 실행
thread.start()
if __name__ == '__main__':
    app.run(debug=True)

    